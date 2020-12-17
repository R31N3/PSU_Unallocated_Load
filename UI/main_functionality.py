import os
import sys
import shutil
from pathlib import Path
from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from PyQt5.QtCore import QRect
from PyQt5.QtWidgets import QApplication, QFileDialog, QMenu, QTableWidgetItem, QMessageBox, QInputDialog
from PyQt5.QtGui import QCursor

from sympy import sympify, SympifyError

from UI.main_ui import MainWindowRefactored
from UI.extended_widgets import *
from UI.enums import ButtonWidgetType


class UnallocatedLoadApplication(MainWindowRefactored):
    def __init__(self):
        super(UnallocatedLoadApplication, self).__init__()
        # Setup Default Values
        self.root_dir = os.getcwd()
        # Setup Variable Values
        self.source_dir = self._select_directory_action(message="Выберите директорию с файлами для работы")
        self.work_dir = self._select_directory_action(message="Выберите рабочую директорию")
        self.result_dir = os.path.join(self.work_dir, "Результаты")
        if not os.path.exists(self.result_dir):
            os.mkdir(self.result_dir)
        # Setup All Sections Actions
        self._setup_files_selection_actions()
        self._setup_work_with_files_actions()

    def _select_directory_action(self, message) -> str:
        directory = QFileDialog.getExistingDirectory(self, message, directory=self.root_dir) or self.root_dir
        return directory

    def _setup_files_selection_actions(self):
        # Setup Files Selection Section Actions
        self.load_files_selection_button.clicked.connect(
            lambda: self._files_selection_button_action(files_list=self.load_files_list)
        )
        self.load_files_list.itemClicked['QListWidgetItem*'].connect(
            lambda: self._deletion_from_files_list_menu(files_list=self.load_files_list)
        )
        self.teachers_files_selection_button.clicked.connect(
            lambda: self._files_selection_button_action(files_list=self.teachers_files_list)
        )
        self.teachers_files_list.itemClicked['QListWidgetItem*'].connect(
            lambda: self._deletion_from_files_list_menu(files_list=self.teachers_files_list)
        )

    @staticmethod
    def _list_fill_action(acting_list: QListWidgetExtended):
        acting_list.clear()
        for selected in acting_list.selected_files:
            acting_list.addItem(selected[0])

    def _files_selection_button_action(
            self, files_list: QListWidgetExtended
    ):
        for selected_file in QFileDialog.getOpenFileNames(
                self, "Выберите файлы", filter="Excel Files (*.xls *.xlsx)", directory=self.source_dir
        )[0]:
            selected = (os.path.basename(selected_file), selected_file)
            if Path(selected_file).suffix in [".xlsx", ".xls"] and selected not in files_list.selected_files:
                files_list.selected_files.append(selected)

        self._list_fill_action(acting_list=files_list)

    def _deletion_from_files_list_menu(self, files_list: QListWidgetExtended):
        menu = QMenu(files_list)
        menu.addAction("Удалить файл из списка")
        menu.triggered.connect(lambda: self._deletion_from_files_list_action(files_list=files_list))
        menu.exec_(QCursor.pos())

    def _deletion_from_files_list_action(self, files_list: QListWidgetExtended):
        selected_file = files_list.currentItem()
        files_list.selected_files.pop(files_list.indexFromItem(selected_file).row())
        files_list.removeItemWidget(selected_file)
        self._list_fill_action(acting_list=files_list)

    def _setup_work_with_files_actions(self):
        # Setup Refresh Buttons Actions
        self.refresh_load_button.clicked.connect(lambda: self._refresh_opened_files_from_list(
            selected_files=self.load_files_list.selected_files,
            files_path=self.work_dir,
            files_tab_widget=self.load_tab
        ))
        self.refresh_teachers_button.clicked.connect(lambda: self._refresh_opened_files_from_list(
            selected_files=self.teachers_files_list.selected_files,
            files_path=self.result_dir,
            files_tab_widget=self.teachers_tab
        ))

        # Setup Table Interaction Buttons
        self.add_load_row_button.clicked.connect(
            lambda: self._add_row_or_col_action(
                files_tab_widget=self.load_tab, action_button=self.add_load_row_button
            )
        )
        self.add_load_column_button.clicked.connect(
            lambda: self._add_row_or_col_action(
                files_tab_widget=self.load_tab, action_button=self.add_load_column_button
            )
        )
        self.add_load_sheet_button.clicked.connect(
            lambda: self._add_tab_action(files_tab_widget=self.load_tab)
        )
        self.add_teachers_row_button.clicked.connect(
            lambda: self._add_row_or_col_action(
                files_tab_widget=self.teachers_tab, action_button=self.add_teachers_row_button
            )
        )
        self.add_teachers_column_button.clicked.connect(
            lambda: self._add_row_or_col_action(
                files_tab_widget=self.teachers_tab, action_button=self.add_teachers_column_button
            )
        )
        self.add_teachers_sheet_button.clicked.connect(
            lambda: self._add_tab_action(files_tab_widget=self.teachers_tab)
        )

        self.end_work_button.clicked.connect(
            lambda: self._end_work_button()
        )

    def _refresh_opened_files_from_list(
            self,
            selected_files: List[str],
            files_path: str,
            files_tab_widget: QTabWidgetExtended
    ):
        for selected_file in selected_files:
            file_name = selected_file[0]
            temp_file_path = os.path.join(files_path, file_name)

            if os.path.exists(temp_file_path) is False or files_tab_widget.opened_files.get(temp_file_path) is None:
                if os.path.exists(temp_file_path) is False:
                    shutil.copyfile(selected_file[1], temp_file_path)
                opened_file = load_workbook(filename=temp_file_path)
                files_tab_widget.opened_files[temp_file_path] = opened_file
            else:
                opened_file = load_workbook(filename=temp_file_path)

            for sheet_name in opened_file.sheetnames:
                tab_name = f"{file_name} -> {sheet_name}"
                if files_tab_widget.opened_files_tables.get(tab_name) is None:
                    new_tab = QTabWidgetTabExtended(object_name=tab_name)
                    files_tab_widget.opened_files_tables[tab_name] = QTableWidgetExtended(
                        geometry=QRect(-1, -1, 605, 485),
                        object_name=tab_name,
                        enable_custom_drag_n_drop=True,
                        parent=new_tab
                    )
                    files_tab_widget.addTab(new_tab, tab_name)
                table_widget = files_tab_widget.opened_files_tables[tab_name]
                self._parse_and_fill_sheet_values_into_table(
                    opened_file=opened_file,
                    file_path=temp_file_path,
                    sheet_name=sheet_name,
                    table_widget=table_widget
                )

    def _parse_and_fill_sheet_values_into_table(
            self,
            opened_file: Workbook,
            file_path: str,
            sheet_name: str,
            table_widget: QTableWidgetExtended
    ):
        worksheet = opened_file[sheet_name]
        for items in worksheet.merged_cell_ranges:
            worksheet.unmerge_cells(str(items))

        headers = [item.value for item in worksheet[8] if item.value is not None]
        row = 1
        rows = []
        table_widget.setColumnCount(len(headers))
        table_widget.setHorizontalHeaderLabels(headers)
        counter = 0
        while worksheet[row][0].value is not None or counter < 5:
            row_values = []
            table_widget.insertRow(table_widget.rowCount())
            for col in range(len(headers)):
                row_values.append(QTableWidgetItem(str(worksheet[row][col].value or "")))
            rows.append(row_values)
            row += 1
            if worksheet[row][0].value is None:
                counter += 1
            else:
                counter = 0

        row = 0
        for row_values in rows:
            for col in range(len(headers)):
                table_widget.setItem(row, col, row_values[col])
            row += 1

        table_widget.resizeColumnsToContents()
        table_widget.itemChanged.connect(lambda: self._change_table_cell_value_action(
            table_widget=table_widget, current_workbook=opened_file, sheet_name=sheet_name, file_path=file_path
        ))

    def _change_table_cell_value_action(
            self, table_widget: QTableWidgetExtended, current_workbook: Workbook, sheet_name: str, file_path: str
    ):
        current_item = table_widget.currentItem()
        if current_item is not None:
            text = self._process_cell_value(current_item.text())

            current_item.setText(text)
            current_worksheet: Worksheet = current_workbook[sheet_name]
            current_worksheet.cell(row=current_item.row() + 1, column=current_item.column() + 1).value = text
            current_workbook.save(file_path)

    @staticmethod
    def _process_cell_value(cell_value: str) -> str:
        value_to_check = cell_value
        expressions = ["/", "+", "*", "^", "-", "**"]
        for expr in expressions:
            value_to_check = value_to_check.replace(expr, "")

        for symbol_to_check in value_to_check:
            if symbol_to_check.isdigit() is False and symbol_to_check not in {" "}:
                return cell_value

        try:
            return str(sympify(cell_value))
        except SympifyError:
            return cell_value

    def _add_row_or_col_action(
            self, files_tab_widget: QTabWidgetExtended, action_button: QPushButtonExtended
    ):
        active_tab = files_tab_widget.currentWidget()
        if active_tab is None:
            self._show_warning(title="Ошибка добавления ячеек", text="Сначала добавьте файл!")
            return

        table_name = active_tab.objectName()
        current_table = files_tab_widget.opened_files_tables[table_name]
        if action_button.button_widget_type == ButtonWidgetType.add_col_button:
            current_table.insertColumn(current_table.columnCount())
        elif action_button.button_widget_type == ButtonWidgetType.add_row_button:
            current_table.insertRow(current_table.rowCount())

    @staticmethod
    def _show_warning(title: str, text: str):
        warning = QMessageBox()
        warning.setWindowTitle(title)
        warning.setText(text)
        warning.setIcon(warning.Warning)
        warning.setStandardButtons(warning.Ok)
        warning.exec()

    def _add_tab_action(self, files_tab_widget: QTabWidgetExtended):
        file_name, result = QInputDialog.getText(self, "Добавление файла", "Введите название файла")
        file_path = None
        if files_tab_widget is self.load_tab:
            file_path = f"{os.path.join(self.work_dir, file_name)}.xlsx"
        else:
            file_path = f"{os.path.join(self.work_dir, 'Результаты', file_name)}.xlsx"
        while file_name is None or files_tab_widget.opened_files.get(file_path) is not None or file_path is None:
            if not result:
                return
            self._show_warning(title="Ошибка добавления файла", text="Название листа пустое или уже существует!")
            file_name, result = QInputDialog.getText(self, "Добавление файла", "Введите название файла")

            if files_tab_widget is self.load_tab:
                file_path = f"{os.path.join(self.work_dir, file_name)}.xlsx"
            else:
                file_path = f"{os.path.join(self.work_dir, 'Результаты', file_name)}.xlsx"

        sheet_name, result = QInputDialog.getText(self, "Добавление листа", "Введите название листа")
        while sheet_name is None or len(sheet_name) == 0:
            if not result:
                return
            self._show_warning(title="Ошибка добавления листа", text="Название листа не может быть пустым")
            sheet_name, result = QInputDialog.getText(self, "Добавление листа", "Введите название листа")

        workbook = Workbook()
        workbook.worksheets[0].title = sheet_name
        workbook.save(file_path)
        files_tab_widget.opened_files[file_path] = load_workbook(file_path)

        tab_name = f"{file_name} -> {sheet_name}"
        new_tab = QTabWidgetTabExtended(object_name=tab_name)
        files_tab_widget.opened_files_tables[tab_name] = QTableWidgetExtended(
            geometry=QRect(-1, -1, 605, 485),
            object_name=tab_name,
            enable_custom_drag_n_drop=True,
            parent=new_tab
        )
        files_tab_widget.addTab(new_tab, tab_name)
        files_tab_widget.setCurrentWidget(new_tab)
        table = files_tab_widget.opened_files_tables[tab_name]
        table.itemChanged.connect(lambda: self._change_table_cell_value_action(
            table_widget=table, current_workbook=workbook, sheet_name=sheet_name, file_path=file_path
        ))

    def _end_work_button(self):
        notification = QMessageBox
        button_clicked = notification.question(
            self, "Завершение обработки",
            "Вы уверены, что хотите завершить обработку?\nРезультаты работы будут храниться в подкаталоге "
            "'Результаты' выбранной рабочей директории",
            notification.Yes | notification.No
        )

        if button_clicked == notification.Yes:
            for file_path, workbook in {**self.load_tab.opened_files, **self.teachers_tab.opened_files}.items():
                workbook.save(file_path)
            self.instance().quit()


def main():
    app = QApplication(sys.argv)
    application_window = UnallocatedLoadApplication()
    application_window.show()
    app.exec_()


if __name__ == '__main__':
    main()
