import sys
import os
import shutil
import openpyxl
import sympy
from typing import Dict, List, Tuple
from pathlib import Path
from PyQt5 import QtWidgets, QtGui, QtCore
from openpyxl.worksheet.worksheet import Worksheet

import UI.Main_Window.main_window as main_window


class ListItem(QtWidgets.QListWidgetItem):
    def __init__(self, visible_text: str):
        super().__init__(visible_text)


class MainWindow(QtWidgets.QMainWindow, main_window.Ui_MainWindow):
    def __init__(self):
        super().__init__()

        # Переменные
        self.root_dir = os.getcwd()

        # selected List[Tuple[str, str]] - Path To File AND Filename to View
        self.selected_load: List[Tuple[str, str]] = []
        self.selected_teachers: List[Tuple[str, str]] = []

        # opened Dict[str, openpyxl.Workbook] - Dict of Opened Files
        self.opened_load: Dict[str, openpyxl.Workbook] = {}
        self.opened_teachers: Dict[str, openpyxl.Workbook] = {}

        # tables Dict[str, QtWidgets.QTableWidget] - Loaded Tables From Excel or Created by User
        self.load_tables: Dict[str, QtWidgets.QTableWidget] = {}
        self.teachers_tables: Dict[str, QtWidgets.QTableWidget] = {}

        # Инициализация
        self.setupUi(self)
        self.source_dir = (
            QtWidgets.QFileDialog.getExistingDirectory(
                self, "Выберите директорию с файлами для работы", directory=self.root_dir
            ) or self.root_dir
        )
        self.work_dir = (
                QtWidgets.QFileDialog.getExistingDirectory(
                    self, "Выберите рабочую директорию", directory=self.root_dir
                ) or self.root_dir
        )
        self.result_dir = os.path.join(self.work_dir, "Результаты")
        if not os.path.exists(self.result_dir):
            os.mkdir(self.result_dir)

        # Действия
        self.LoadSelection.clicked.connect(lambda: self.selection_button(
            qlist=self.SelectedLoad,
            data_list=self.selected_load
        ))
        self.SelectedLoad.itemClicked['QListWidgetItem*'].connect(lambda: self.deletion_menu(
            data_list=self.selected_load,
            qlist=self.SelectedLoad,
        ))
        self.RefreshLoad.clicked.connect(lambda: self.refresh(
            data_list=self.selected_load,
            opened_files=self.opened_load,
            tables=self.load_tables,
            tab_widget=self.LoadTabs
        ))
        self.AddLoadRow.clicked.connect(lambda: self.add_row_button(
            tab_widget=self.LoadTabs, tables=self.load_tables
        ))
        self.AddLoadColumn.clicked.connect(lambda: self.add_column_button(
            tab_widget=self.LoadTabs, tables=self.load_tables
        ))
        self.AddLoadTab.clicked.connect(lambda: self.add_tab_button(
            tab_widget=self.LoadTabs, tables=self.load_tables, opened_files=self.opened_load
        ))

        self.TeachersSelection.clicked.connect(lambda: self.selection_button(
            qlist=self.SelectedTeachers,
            data_list=self.selected_teachers
        ))
        self.SelectedTeachers.itemClicked['QListWidgetItem*'].connect(lambda: self.deletion_menu(
            data_list=self.selected_teachers,
            qlist=self.SelectedTeachers,
        ))
        self.RefreshTeachers.clicked.connect(lambda: self.refresh(
            data_list=self.selected_teachers,
            opened_files=self.opened_teachers,
            tables=self.teachers_tables,
            tab_widget=self.TeachersTabs
        ))
        self.AddTeachersRow.clicked.connect(
            lambda: self.add_row_button(tab_widget=self.TeachersTabs, tables=self.teachers_tables)
        )
        self.AddTeachersColumn.clicked.connect(
            lambda: self.add_column_button(tab_widget=self.TeachersTabs, tables=self.teachers_tables)
        )
        self.AddTeachersTab.clicked.connect(
            lambda: self.add_tab_button(
                tab_widget=self.TeachersTabs, tables=self.teachers_tables, opened_files=self.opened_teachers
            )
        )

        self.EndWork.clicked.connect(self.end_work_button)

    @staticmethod
    def fill_list(qlist: QtWidgets.QListWidget, selected_files):
        qlist.clear()
        for selected in selected_files:
            qlist.addItem(selected[0])

    def selection_button(self, data_list: list, qlist: QtWidgets.QListWidget):
        for selected_file in QtWidgets.QFileDialog.getOpenFileNames(
                self, "Выберите файлы", filter="Excel Files (*.xls *.xlsx)", directory=self.source_dir
        )[0]:
            selected = (os.path.basename(selected_file), selected_file)
            if Path(selected_file).suffix in [".xlsx", ".xls"] and selected not in data_list:
                data_list.append(selected)

        self.fill_list(qlist=qlist, selected_files=data_list)

    def deletion_from_list(self, data_list: list, qlist: QtWidgets.QListWidget):
        data_list.pop(qlist.indexFromItem(qlist.currentItem()).row())
        qlist.removeItemWidget(qlist.currentItem())
        self.fill_list(qlist=qlist, selected_files=data_list)

    def deletion_menu(self, data_list: list, qlist: QtWidgets.QListWidget):
        menu = QtWidgets.QMenu(qlist)
        menu.addAction("Удалить файл из списка")
        menu.triggered.connect(lambda: self.deletion_from_list(data_list=data_list, qlist=qlist))
        menu.exec_(QtGui.QCursor.pos())

    def refresh(
            self, data_list: list, opened_files: Dict[str, openpyxl.Workbook],
            tab_widget: QtWidgets.QTabWidget, tables: Dict[str, QtWidgets.QTableWidget]
    ):
        for load_file in data_list:
            file_name = load_file[0]
            if tab_widget.objectName() == self.LoadTabs.objectName():
                target_temp_file = os.path.join(self.work_dir, file_name)
            else:
                target_temp_file = os.path.join(self.work_dir, "Результаты", file_name)

            if not os.path.exists(target_temp_file):
                shutil.copyfile(load_file[1], target_temp_file)
                opened_file = openpyxl.load_workbook(filename=target_temp_file)
                opened_files[target_temp_file] = opened_file
            elif not opened_files.get(target_temp_file):
                opened_file = openpyxl.load_workbook(filename=target_temp_file)
                opened_files[target_temp_file] = opened_file
            else:
                opened_file = openpyxl.load_workbook(filename=target_temp_file)

            for sheet_name in opened_file.sheetnames:
                tab_name = f"{file_name} -> {sheet_name}"
                if not tables.get(tab_name):
                    tab = QtWidgets.QWidget()
                    tab.setObjectName(tab_name)
                    tables[tab_name] = QtWidgets.QTableWidget(tab)
                    tables[tab_name].setGeometry(QtCore.QRect(-10, -10, 861, 901))
                    tables[tab_name].setObjectName(tab_name)
                    tables[tab_name].setColumnCount(0)
                    tables[tab_name].setRowCount(0)
                    tab_widget.addTab(tab, tab_name)

                table = tables[tab_name]

                ws = opened_file[sheet_name]

                for items in ws.merged_cell_ranges:
                    ws.unmerge_cells(str(items))

                headers = [item.value for item in ws[8] if item.value is not None]
                row = 1
                rows = []
                table.setColumnCount(len(headers))
                table.setHorizontalHeaderLabels(headers)
                counter = 0
                while ws[row][0].value is not None or counter < 5:
                    row_values = []
                    table.insertRow(table.rowCount())
                    for col in range(len(headers)):
                        row_values.append(QtWidgets.QTableWidgetItem(str(ws[row][col].value or "")))
                    rows.append(row_values)
                    row += 1
                    if ws[row][0].value is None:
                        counter += 1
                    else:
                        counter = 0

                row = 0
                for row_values in rows:
                    for col in range(len(headers)):
                        table.setItem(row, col, row_values[col])
                    row += 1

                table.resizeColumnsToContents()
                table.itemChanged.connect(lambda: self.change_table_cell_value(
                    active_table=table, current_workbook=opened_file, sheet_name=sheet_name, file_path=target_temp_file
                ))

    @staticmethod
    def show_warning(title: str, text: str):
        warning = QtWidgets.QMessageBox()
        warning.setWindowTitle(title)
        warning.setText(text)
        warning.setIcon(warning.Warning)
        warning.setStandardButtons(warning.Ok)
        warning.exec()

    def add_tab_button(
            self, tab_widget: QtWidgets.QTabWidget, tables: Dict[str, QtWidgets.QTableWidget],
            opened_files: Dict[str, openpyxl.Workbook]
    ):
        file_name, result = QtWidgets.QInputDialog.getText(self, "Добавление файла", "Введите название файла")
        if tab_widget.objectName() == self.LoadTabs.objectName():
            file_path = f"{os.path.join(self.work_dir, file_name)}.xlsx"
        else:
            file_path = f"{os.path.join(self.work_dir, 'Результаты', file_name)}.xlsx"
        while not file_name or opened_files.get(file_path) is not None:
            if not result:
                return
            self.show_warning(title="Ошибка добавления файла", text="Название листа пустое или уже существует!")
            file_name, result = QtWidgets.QInputDialog.getText(self, "Добавление файла", "Введите название файла")

            if tab_widget.objectName() == self.LoadTabs.objectName():
                file_path = f"{os.path.join(self.work_dir, file_name)}.xlsx"
            else:
                file_path = f"{os.path.join(self.work_dir, 'Результаты', file_name)}.xlsx"

        sheet_name, result = QtWidgets.QInputDialog.getText(self, "Добавление листа", "Введите название листа")
        while not sheet_name:
            if not result:
                return
            self.show_warning(title="Ошибка добавления листа", text="Название листа не может быть пустым")
            sheet_name, result = QtWidgets.QInputDialog.getText(self, "Добавление листа", "Введите название листа")

        workbook = openpyxl.Workbook()
        workbook.worksheets[0].title = sheet_name
        workbook.save(file_path)
        opened_files[file_path] = openpyxl.load_workbook(file_path)

        tab_name = f"{file_name} -> {sheet_name}"
        tab_to_create = QtWidgets.QWidget()
        tab_to_create.setObjectName(tab_name)
        tables[tab_name] = QtWidgets.QTableWidget(tab_to_create)
        tables[tab_name].setGeometry(QtCore.QRect(-10, -10, 861, 901))
        tables[tab_name].setObjectName(tab_name)
        tables[tab_name].setColumnCount(0)
        tables[tab_name].setRowCount(0)
        tab_widget.addTab(tab_to_create, tab_name)
        tab_widget.setCurrentWidget(tab_to_create)
        table = tables[tab_name]
        table.itemChanged.connect(lambda: self.change_table_cell_value(
            active_table=table, current_workbook=opened_files[file_path], sheet_name=sheet_name, file_path=file_path
        ))

    def add_row_button(self, tab_widget: QtWidgets.QTabWidget, tables: Dict[str, QtWidgets.QTableWidget]):
        active_tab = tab_widget.currentWidget()
        if not active_tab:
            self.show_warning(title="Ошибка добавления ячеек", text="Сначала добавьте файл!")
            return

        table_name = active_tab.objectName()
        current_table = tables[table_name]
        current_table.insertRow(current_table.rowCount())

    def add_column_button(self, tab_widget: QtWidgets.QTabWidget, tables: Dict[str, QtWidgets.QTableWidget]):
        active_tab = tab_widget.currentWidget()
        if not active_tab:
            self.show_warning(title="Ошибка добавления ячеек", text="Сначала добавьте файл!")
            return

        table_name = active_tab.objectName()
        current_table = tables[table_name]
        current_table.insertColumn(current_table.columnCount())

    @staticmethod
    def process_cell_value(cell_value: str):
        value_to_check = cell_value
        expressions = ["/", "+", "*", "^", "-", "**"]
        for expr in expressions:
            value_to_check = value_to_check.replace(expr, "")

        for symbol_to_check in value_to_check:
            if symbol_to_check.isdigit() is False and symbol_to_check not in {" "}:
                return cell_value

        try:
            return str(sympy.sympify(cell_value))
        except sympy.SympifyError:
            return cell_value

    def change_table_cell_value(
            self, active_table: QtWidgets.QTableWidget, current_workbook: openpyxl.Workbook,
            sheet_name: str, file_path: str
    ):
        current_item = active_table.currentItem()

        text = self.process_cell_value(current_item.text())

        current_item.setText(text)
        current_worksheet: Worksheet = current_workbook[sheet_name]
        current_worksheet.cell(row=current_item.row() + 1, column=current_item.column() + 1).value = text
        current_workbook.save(file_path)

    def end_work_button(self):
        notification = QtWidgets.QMessageBox
        button_clicked = notification.question(
            self, "Завершение обработки",
            "Вы уверены, что хотите завершить обработку?\nРезультаты работы будут храниться в подкаталоге "
            "'Результаты' выбранной рабочей директории",
            notification.Yes | notification.No
        )

        if button_clicked == notification.Yes:
            for file_path, workbook in {**self.opened_teachers, **self.opened_load}.items():
                workbook.save(file_path)
            QtCore.QCoreApplication.instance().quit()


def main():
    app = QtWidgets.QApplication(sys.argv)
    application_window = MainWindow()
    application_window.show()
    app.exec_()


if __name__ == '__main__':
    main()
